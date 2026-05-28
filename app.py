"""
공연 티켓 예매 시스템 - Flask 백엔드
삼육대학교 선교 70주년 기념관 대강당 (총 2,352석)
"""
from flask import Flask, request, session, jsonify, render_template, send_file
import sqlite3, hashlib, os, datetime, io, json, re, secrets
import threading, time, uuid
from collections import OrderedDict
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sda-ticket-secret-key-2026-sjhc-CHANGE-IN-PROD')
# ── 세션 쿠키 최적화 (300명 동시접속 보안·성능) ─────────────────
app.config['SESSION_COOKIE_HTTPONLY']   = True     # XSS: JS에서 쿠키 접근 차단
app.config['SESSION_COOKIE_SAMESITE']   = 'Lax'   # CSRF 방어 + 일반 탐색 허용
app.config['SESSION_COOKIE_NAME']       = 'ts'    # 짧은 쿠키명 → 헤더 크기 절감
app.config['PERMANENT_SESSION_LIFETIME'] = 7200   # 세션 유효기간 2시간(초)
app.config['JSON_AS_ASCII']             = False   # 한글 JSON 인코딩 최적화

# ── Flask-Limiter: 로그인 무차별 대입 공격 방어 ─────────────────────
# - 메모리 저장소 사용 (Redis 불필요, 서버 재시작 시 카운터 초기화)
# - 기본 제한 없음 (로그인 라우트에만 @limiter.limit 으로 개별 적용)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,  # IP 기준 카운팅
    default_limits=[],            # 전역 제한 없음 — 로그인에만 적용
    storage_uri='memory://',
)

# ══════════════════════════════════════════════════════════════════
# 대기열 시스템 (Queue System)
# ══════════════════════════════════════════════════════════════════
MAX_ACTIVE      = 20    # 동시 입장 최대 인원 (DB config: max_active)
ACTIVE_TIMEOUT  = 180   # 입장 후 예매 제한시간 초 (DB config: booking_duration)
HEARTBEAT_LIMIT = 90    # 90초 무응답 → 자동 퇴장
SEAT_LOCK_SECS  = 180   # 좌석 선점 시간 = booking_duration 과 동기화

_ql      = threading.Lock()
_active  = {}            # token → {uid, name, entered_at, heartbeat}
_waiting = OrderedDict() # token → {uid, name, joined_at}

# Perf-③: 동시 예매 확정 요청 제한 (SQLite BEGIN IMMEDIATE 충돌 최소화)
_reservation_sem = threading.Semaphore(20)  # 동시 처리 최대 20개 (스레드 32 기준)

def _queue_fill():
    """빈 슬롯에 대기자 입장 (lock 보유 상태에서 호출)"""
    while len(_active) < MAX_ACTIVE and _waiting:
        tok, info = next(iter(_waiting.items()))
        del _waiting[tok]
        _active[tok] = {**info, 'entered_at': time.time(), 'heartbeat': time.time()}

# Perf-②: _ql 보유 중 DB I/O 제거 → 잠금 점유 시간 최소화
def _release_seat_lock_db(token):
    """좌석 lock DB 해제 (_ql 외부에서 호출해야 함)"""
    try:
        _db = os.environ.get("TICKET_DB", os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "tickets.db"))
        conn = sqlite3.connect(_db, timeout=10)
        conn.execute('PRAGMA journal_mode = WAL')
        conn.execute(
            "UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL "
            "WHERE locked_by=? AND status='locked'", (token,))
        conn.commit()
        conn.close()
    except Exception:
        pass

def _queue_kick(token):
    """활성 세션 메모리 퇴장 + 슬롯 보충 (_ql 보유 상태에서 호출)
    DB 좌석 해제는 _ql 해제 후 _release_seat_lock_db(token) 로 처리.
    """
    _active.pop(token, None)
    _queue_fill()

def _queue_cleanup_loop():
    """백그라운드: 하트비트 만료 + 좌석 lock 만료 처리"""
    while True:
        time.sleep(5)
        now = time.time()
        expired_tokens = []
        with _ql:
            # 하트비트 타임아웃
            expired = [t for t, v in list(_active.items())
                       if now - v['heartbeat'] > HEARTBEAT_LIMIT]
            for t in expired:
                _queue_kick(t)          # 메모리만 (_ql 내부)
                expired_tokens.append(t)
            # 30분 초과 대기자 제거
            old = [t for t, v in list(_waiting.items())
                   if now - v['joined_at'] > 1800]
            for t in old:
                _waiting.pop(t, None)
        # Perf-②: _ql 해제 후 DB 작업 (잠금 점유 최소화)
        for t in expired_tokens:
            _release_seat_lock_db(t)
        # DB 좌석 lock 만료 (5분 초과)
        try:
            db_path = os.environ.get("TICKET_DB", os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "tickets.db"))
            conn = sqlite3.connect(db_path, timeout=10)
            conn.execute('PRAGMA journal_mode = WAL')
            expiry = datetime.datetime.now() - datetime.timedelta(seconds=SEAT_LOCK_SECS)
            conn.execute("UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL "
                         "WHERE status='locked' AND locked_at < ?",
                         (expiry.strftime('%Y-%m-%d %H:%M:%S'),))
            conn.commit(); conn.close()
        except Exception: pass

threading.Thread(target=_queue_cleanup_loop, daemon=True).start()
# ══════════════════════════════════════════════════════════════════
# DB 파일: 앱 파일과 같은 디렉토리에 저장
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.environ.get("TICKET_DB", os.path.join(BASE_DIR, "tickets.db"))

# ──────────────────────────────────────────────
# 좌석 구성 정의 (선교 70주년 기념관 대강당 좌석배치도 PDF 기반)
# ──────────────────────────────────────────────
# 1층: 가/라=부채꼴(앞 짧고 뒤 넓음), 나/다=중간 구역
# row_seats 검증:
#  가/라: 6+9+12+13+13 + 14*22 = 53+308 = 361 ✓
#  나/다: 8+10 + 13*20 + 12*5 = 18+260+60 = 338 ✓
SECTIONS = {
    # ── 1층 ──────────────────────────────────────────────────────────────
    # 좌석표(최종).xlsx 실측 기반 정확 배치
    # 가/라: 부채꼴 외곽 구역 (rows=26, 361석)
    #   row_seats 검증: 11*2 + 12*5 + 13*4 + 14*4 + 15*5 + 16*6
    #                 = 22+60+52+56+75+96 = 361 ✓
    # 나/다: 중앙 구역 (rows=26, 338석)
    #   row_seats 검증: 13*26 = 338 ✓
    # disability_cols: 제일 앞 열 장애인 좌석 수 (번호 없이 아이콘만 표시)
    '가': {'floor': 1, 'total': 361, 'label': '가 01~361', 'side': 'left',
           'rows': 26,
           'row_seats': [11,11, 12,12,12,12,12, 13,13,13,13,
                         14,14,14,14, 15,15,15,15,15, 16,16,16,16,16,16],
           'align': 'right',           # 부채꼴 — 내측(무대쪽)이 오른쪽
           'disability_cols': 6,       # 앞열 장애인석 6자리
           'aisle_after_row': 11},     # 11행과 12행 사이 가로 복도
    '나': {'floor': 1, 'total': 338, 'label': '나 01~338', 'side': 'center-left',
           'rows': 26,
           'row_seats': [13]*26,
           'align': 'center',
           'disability_cols': 8,       # 앞열 장애인석 8자리
           'aisle_after_row': 11},
    '다': {'floor': 1, 'total': 338, 'label': '다 01~338', 'side': 'center-right',
           'rows': 26,
           'row_seats': [13]*26,
           'align': 'center',
           'disability_cols': 8,
           'aisle_after_row': 11},
    '라': {'floor': 1, 'total': 361, 'label': '라 01~361', 'side': 'right',
           'rows': 26,
           'row_seats': [11,11, 12,12,12,12,12, 13,13,13,13,
                         14,14,14,14, 15,15,15,15,15, 16,16,16,16,16,16],
           'align': 'left',            # 부채꼴 — 내측(무대쪽)이 왼쪽
           'disability_cols': 6,
           'aisle_after_row': 11},
    # ── 2층 ──────────────────────────────────────────────────────────────
    # 마/바: 측면 발코니 (rows=15, 29석)  2*14+1 = 29 ✓
    # 사: 4+5+6 + 7*13 = 106 ✓   아/카: 10*16 = 160 ✓
    # 자/차: 12*15 = 180 ✓        타: 10*11 = 110 ✓
    '마': {'floor': 2, 'total': 29,  'label': '마 01~29',  'side': 'far-left',
           'rows': 15, 'row_seats': [2]*14 + [1],
           'align': 'center'},
    # 사: 좌외곽 — 상단 좁고 하단 넓은 부채꼴 (좌기준 정렬)
    # row_seats 검증: 5*4 + 6*5 + 7*8 = 20+30+56 = 106 ✓
    '사': {'floor': 2, 'total': 106, 'label': '사 01~106', 'side': 'left',
           'rows': 17,
           'row_seats': [5,5,5,5, 6,6,6,6,6, 7,7,7,7,7,7,7,7],
           'align': 'left'},
    # 아: 중앙좌 — 중간 rows 9-12 출구로 7석, 나머지 12석
    # row_seats 검증: 12*8 + 7*4 + 12*3 = 96+28+36 = 160 ✓
    '아': {'floor': 2, 'total': 160, 'label': '아 01~160', 'side': 'center-left',
           'rows': 15,
           'row_seats': [12,12,12,12,12,12,12,12, 7,7,7,7, 12,12,12],
           'align': 'left'},
    # 자: 중앙 — rows 1-2 CONTROL BOOTH(blocked), rows 3-15 정상 좌석
    # row_seats 검증: 12*15 = 180 ✓  (rows 1-2 = 24석 blocked → BLOCKED_SEATS 처리)
    '자': {'floor': 2, 'total': 180, 'label': '자 01~180', 'side': 'center',
           'rows': 15, 'row_seats': [12]*15,
           'align': 'center'},
    # 차: 중앙우 — 전 행 12석 균일
    # row_seats 검증: 12*15 = 180 ✓
    '차': {'floor': 2, 'total': 180, 'label': '차 01~180', 'side': 'center-right',
           'rows': 15, 'row_seats': [12]*15,
           'align': 'center'},
    # 카: 중앙우 — 아구역 대칭 (출구가 좌측, 우기준 정렬)
    # row_seats 검증: 12*8 + 7*4 + 12*3 = 96+28+36 = 160 ✓
    '카': {'floor': 2, 'total': 160, 'label': '카 01~160', 'side': 'right',
           'rows': 15,
           'row_seats': [12,12,12,12,12,12,12,12, 7,7,7,7, 12,12,12],
           'align': 'right'},
    # 타: 우외곽 — 사구역 대칭 (우기준 정렬), 마지막 행 4석 추가
    # row_seats 검증: 5*4 + 6*5 + 7*8 + 4*1 = 20+30+56+4 = 110 ✓
    '타': {'floor': 2, 'total': 110, 'label': '타 01~110', 'side': 'far-right',
           'rows': 18,
           'row_seats': [5,5,5,5, 6,6,6,6,6, 7,7,7,7,7,7,7,7, 4],
           'align': 'right'},
    '바': {'floor': 2, 'total': 29,  'label': '바 01~29',  'side': 'far-right-balcony',
           'rows': 15, 'row_seats': [2]*14 + [1],
           'align': 'center'},
}

# 좌석 배치도에서 완전 제외할 위치 (section, row, col)
# — DB에 삽입하지 않으며, 통계·차단 목록에서도 제외됨
# — SVG에는 renderFloor2()의 CONTROL BOOTH 오버레이로 별도 표시
EXCLUDED_SEATS = {
    # 자 구역: rows 1-2 전체 = CONTROL BOOTH 설치 위치 (24개 위치, 좌석 아님)
    '자': set((r, c) for r in (1, 2) for c in range(1, 13)),
}
# 하위 호환: 구버전 변수명 유지 (init_db 등에서 참조)
BLOCKED_SEATS = EXCLUDED_SEATS

# DB 좌석 버전 - 배치 변경 시 숫자를 올리면 seats/reservations 자동 초기화
SEAT_VERSION = '5.0'   # 버전 변경 시 seats/reservations 자동 초기화

# ──────────────────────────────────────────────
# DB 유틸리티
# ──────────────────────────────────────────────
def get_db():
    # timeout=15: 동시 쓰기 경합 시 대기 시간 (300명 기준 상향)
    conn = sqlite3.connect(DATABASE, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    # WAL 모드: 읽기와 쓰기가 서로 차단하지 않음 (300명 동시접속)
    conn.execute('PRAGMA journal_mode = WAL')
    # 동기화 수준 완화: 쓰기 성능 향상 (WAL 모드에서 안전)
    conn.execute('PRAGMA synchronous = NORMAL')
    # 페이지 캐시 32MB (기본 ~2MB) — 반복 읽기 DB I/O 대폭 감소
    conn.execute('PRAGMA cache_size = -32000')
    # 메모리맵 I/O 128MB — WAL 읽기 성능 향상
    conn.execute('PRAGMA mmap_size = 134217728')
    # WAL 체크포인트: 2000페이지마다 (기본 1000, 쓰기 빈도 분산)
    conn.execute('PRAGMA wal_autocheckpoint = 2000')
    return conn

# ── Perf-① config 인-메모리 캐시 (TTL 5초) ───────────────────
_config_cache      = {}              # key → (value, timestamp)
_config_cache_lock = threading.Lock()
_CONFIG_TTL        = 5               # 초 단위

def _invalidate_config_cache(key=None):
    """관리자 설정 변경 시 캐시 즉시 무효화"""
    with _config_cache_lock:
        if key:
            _config_cache.pop(key, None)
        else:
            _config_cache.clear()

def get_config_value(key, default=None):
    """config 테이블에서 설정값 조회 (TTL 캐시 적용)"""
    now = time.time()
    with _config_cache_lock:
        if key in _config_cache:
            val, ts = _config_cache[key]
            if now - ts < _CONFIG_TTL:
                return val if val is not None else default
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM config WHERE key=?", (key,)).fetchone()
        conn.close()
        val = row[0] if (row and row[0] is not None and row[0] != '') else None
        with _config_cache_lock:
            _config_cache[key] = (val, time.time())
        return val if val is not None else default
    except Exception:
        return default

_init_db_lock = threading.Lock()
_init_db_done = False

def init_db():
    """DB 초기화 — 멀티스레드/멀티프로세스 환경에서 1회만 실행되도록 보호"""
    global _init_db_done
    # 이미 초기화 완료된 경우 즉시 반환 (lock 없이)
    if _init_db_done:
        return
    with _init_db_lock:
        # double-checked locking: lock 획득 후 재확인
        if _init_db_done:
            return
        _run_init_db()
        _init_db_done = True

def _run_init_db():
    conn = get_db()
    c = conn.cursor()

    # 회원 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        username  TEXT UNIQUE NOT NULL,
        password  TEXT NOT NULL,
        name      TEXT NOT NULL,
        phone     TEXT NOT NULL,
        church    TEXT NOT NULL,
        position  TEXT NOT NULL,
        role      TEXT DEFAULT 'user',
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )''')

    # 좌석 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS seats (
        id        TEXT PRIMARY KEY,
        floor     INTEGER NOT NULL,
        section   TEXT NOT NULL,
        row_num   INTEGER NOT NULL,
        col_num   INTEGER NOT NULL,
        seat_no   TEXT NOT NULL,
        status    TEXT DEFAULT 'available'
    )''')

    # 예매 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS reservations (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        res_code       TEXT UNIQUE NOT NULL,
        user_id        INTEGER NOT NULL,
        seat_id        TEXT NOT NULL,
        status         TEXT DEFAULT 'confirmed',
        reserved_at    TEXT DEFAULT (datetime('now','localtime')),
        cancelled_at   TEXT,
        memo           TEXT,
        payment_status TEXT DEFAULT NULL,
        payment_at     TEXT DEFAULT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(seat_id) REFERENCES seats(id)
    )''')

    # 기존 DB 마이그레이션: payment 컬럼 추가
    for col, definition in [('payment_status', 'TEXT DEFAULT NULL'),
                             ('payment_at',     'TEXT DEFAULT NULL')]:
        try:
            c.execute(f'ALTER TABLE reservations ADD COLUMN {col} {definition}')
            conn.commit()
        except Exception:
            pass  # 이미 컬럼이 존재하면 무시

    # 기존 DB 마이그레이션: 환불 계좌 컬럼 추가 (users 테이블)
    for col, definition in [('refund_bank',    'TEXT DEFAULT NULL'),
                             ('refund_account', 'TEXT DEFAULT NULL'),
                             ('refund_holder',  'TEXT DEFAULT NULL')]:
        try:
            c.execute(f'ALTER TABLE users ADD COLUMN {col} {definition}')
            conn.commit()
        except Exception:
            pass  # 이미 컬럼이 존재하면 무시

    # 기존 DB 마이그레이션: 좌석 선점 Lock 컬럼 추가 (seats 테이블)
    for col, definition in [('locked_by', 'TEXT DEFAULT NULL'),
                             ('locked_at', 'TEXT DEFAULT NULL')]:
        try:
            c.execute(f'ALTER TABLE seats ADD COLUMN {col} {definition}')
            conn.commit()
        except Exception:
            pass

    # CONTROL BOOTH 좌석 완전 제거 마이그레이션
    # (자 구역 rows 1-2 = 24개 위치는 좌석이 아닌 방송 부스 — DB에서 완전 삭제)
    cb_del = c.execute(
        "DELETE FROM seats WHERE section='자' AND row_num IN (1, 2)"
    ).rowcount
    if cb_del:
        print(f"CONTROL BOOTH 위치 {cb_del}개 seats 테이블에서 제거")
        conn.commit()

    # 데이터 정합성 복구: 확정 예매가 있는 좌석 중 status가 'available'/'locked'인 것을 'reserved'로 동기화
    fixed = c.execute('''
        UPDATE seats SET status='reserved', locked_by=NULL, locked_at=NULL
        WHERE id IN (
            SELECT seat_id FROM reservations WHERE status='confirmed'
        ) AND status IN ('available', 'locked')
    ''').rowcount
    if fixed:
        print(f"좌석 상태 복구: {fixed}석 (available/locked → reserved 동기화)")
    conn.commit()

    # 관리자 계정 생성
    admin_pw = hash_password('Admin1234!')
    c.execute('''INSERT OR IGNORE INTO users
        (username, password, name, phone, church, position, role)
        VALUES (?,?,?,?,?,?,?)''',
        ('admin', admin_pw, '관리자', '02-3399-4051', '서중한합회', '목사', 'admin'))
    # 관리자 정보 항상 최신 유지 (전화번호 등 변경 시 자동 반영)
    c.execute('''UPDATE users SET phone=?, church=?, position=?
        WHERE username='admin' ''',
        ('02-3399-4051', '서중한합회', '목사'))
    conn.commit()

    # ── 관리자 비밀번호 해시 형식 마이그레이션 (SHA256 → PBKDF2) ──
    admin_row = c.execute(
        "SELECT id, password FROM users WHERE username='admin'"
    ).fetchone()
    if admin_row and _SHA256_RE.match(admin_row['password']):
        c.execute('UPDATE users SET password=? WHERE id=?',
                  (hash_password('Admin1234!'), admin_row['id']))
        conn.commit()
        print("관리자 비밀번호 해시를 PBKDF2-SHA256 형식으로 업그레이드했습니다.")

    # ── DB 버전 체크 (좌석 배치 변경 시 자동 초기화) ──
    c.execute('''CREATE TABLE IF NOT EXISTS config
        (key TEXT PRIMARY KEY, value TEXT)''')
    # 기본 설정값 초기화 (없을 때만)
    c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('max_active', '20')")
    c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('booking_open_time', '')")
    c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('booking_duration', '180')")
    c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('ticket_price_floor1', '10000')")
    c.execute("INSERT OR IGNORE INTO config (key, value) VALUES ('ticket_price_floor2', '10000')")
    conn.commit()

    ver_row = c.execute("SELECT value FROM config WHERE key='seat_version'").fetchone()
    cur_ver = ver_row[0] if ver_row else None

    if cur_ver != SEAT_VERSION:
        print(f"좌석 버전 변경 ({cur_ver} -> {SEAT_VERSION}): seats/reservations 초기화")
        # 데이터 손실 방지: 삭제 전 백업 테이블 보존
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        try:
            c.execute(f"CREATE TABLE IF NOT EXISTS reservations_backup_{ts} AS SELECT * FROM reservations")
            c.execute(f"CREATE TABLE IF NOT EXISTS seats_backup_{ts} AS SELECT * FROM seats")
            print(f"  → 백업 완료: reservations_backup_{ts}, seats_backup_{ts}")
        except Exception as e:
            print(f"  → 백업 실패 (무시): {e}")
        c.execute('DELETE FROM reservations')
        c.execute('DELETE FROM seats')
        c.execute("INSERT OR REPLACE INTO config VALUES ('seat_version', ?)", (SEAT_VERSION,))
        conn.commit()

    # 좌석 데이터 초기화
    existing = c.execute('SELECT COUNT(*) FROM seats').fetchone()[0]
    if existing == 0:
        seat_rows = []
        # EXCLUDED_SEATS: DB에 삽입하지 않을 위치 집합 (CONTROL BOOTH 등)
        excl = {sec: positions for sec, positions in EXCLUDED_SEATS.items()}

        for sec, cfg in SECTIONS.items():
            seat_num = 1
            exc_set = excl.get(sec, set())
            for r_idx, count in enumerate(cfg['row_seats']):
                for c_idx in range(count):
                    row_1 = r_idx + 1
                    col_1 = c_idx + 1
                    # 완전 제외 좌석 → 삽입하지 않음 (seat_num도 건너뜀)
                    if (row_1, col_1) in exc_set:
                        continue
                    seat_rows.append((
                        f"{sec}{seat_num:03d}",
                        cfg['floor'], sec,
                        row_1, col_1,
                        f"{seat_num:03d}", 'available'
                    ))
                    seat_num += 1

        c.executemany('''INSERT INTO seats
            (id, floor, section, row_num, col_num, seat_no, status)
            VALUES (?,?,?,?,?,?,?)''', seat_rows)
        conn.commit()
        excl_cnt = sum(len(v) for v in EXCLUDED_SEATS.values())
        print(f"좌석 {len(seat_rows)}석 초기화 완료 (CONTROL BOOTH 등 {excl_cnt}개 위치 제외됨)")

    conn.close()

# ── 비밀번호 유틸리티 ────────────────────────────────
# 구버전 SHA-256 단순 해시 감지용 (마이그레이션 전용)
_SHA256_RE = re.compile(r'^[a-f0-9]{64}$')

def hash_password(pw):
    """비밀번호 해시 생성 (PBKDF2-SHA256 + salt 16byte)"""
    return generate_password_hash(pw, method='pbkdf2:sha256', salt_length=16)

def verify_password(stored_hash, pw):
    """비밀번호 검증 — 구 SHA256 단순 해시도 자동 감지·처리"""
    if _SHA256_RE.match(stored_hash):
        # 구버전 해시: 평문 SHA256으로 비교
        return stored_hash == hashlib.sha256(pw.encode()).hexdigest()
    return check_password_hash(stored_hash, pw)

# ──────────────────────────────────────────────
# CSRF 보호
# ──────────────────────────────────────────────
def _get_csrf_token():
    """세션에 CSRF 토큰이 없으면 생성 후 반환"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def csrf_protect(f):
    """POST/PUT/DELETE 요청에 대해 X-CSRF-Token 헤더를 검증하는 데코레이터"""
    @wraps(f)
    def decorated(*args, **kwargs):
        # sendBeacon 방식의 queue/leave는 헤더 추가 불가 → CSRF 면제 (해악 없는 정리 작업)
        if request.path == '/api/queue/leave':
            return f(*args, **kwargs)
        if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
            token = request.headers.get('X-CSRF-Token', '')
            expected = session.get('csrf_token', '')
            if not expected or not secrets.compare_digest(token, expected):
                return jsonify({'error': 'CSRF 토큰이 유효하지 않습니다.'}), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/api/csrf-token', methods=['GET'])
def get_csrf_token():
    """클라이언트가 페이지 로드 시 CSRF 토큰을 가져가는 엔드포인트"""
    return jsonify({'csrf_token': _get_csrf_token()})

# ──────────────────────────────────────────────
# 데코레이터
# ──────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': '로그인이 필요합니다.'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': '관리자 권한이 필요합니다.'}), 403
        return f(*args, **kwargs)
    return decorated

def gen_res_code(conn=None):
    """예매번호 생성 — 암호학적 난수(secrets) 사용, 중복 재시도 최대 20회
    conn 인자를 전달하면 해당 연결을 재사용 (BEGIN IMMEDIATE 트랜잭션 내부용)
    """
    year = datetime.datetime.now().strftime('%Y')
    ext_conn = conn is not None
    if not ext_conn:
        conn = get_db()
    try:
        for _ in range(20):
            seq  = secrets.randbelow(9000) + 1000   # 1000~9999 (암호학적 난수)
            code = f'RES-W{year}{seq}'
            dup  = conn.execute(
                'SELECT id FROM reservations WHERE res_code=?', (code,)
            ).fetchone()
            if not dup:
                return code
        # 극히 드문 충돌: 6자리로 확장
        return f'RES-W{year}{secrets.randbelow(900000) + 100000}'
    finally:
        if not ext_conn:
            conn.close()

# ──────────────────────────────────────────────
# 페이지 라우트
# ──────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/register')
def register_page():
    return render_template('register.html')

@app.route('/booking')
def booking_page():
    if 'user_id' not in session:
        # 로그인 후 메인화면으로 → 대기열을 통해 정상 입장
        return render_template('login.html', redirect='/')
    return render_template('booking.html')

@app.route('/queue')
def queue_page():
    if 'user_id' not in session:
        return render_template('login.html', redirect='/queue')
    return render_template('queue.html')

# ──────────────────────────────────────────────
# API - 대기열 (Queue)
# ──────────────────────────────────────────────
@app.route('/api/queue/join', methods=['POST'])
@csrf_protect
@login_required
def queue_join():
    """대기열 참가 또는 즉시 입장"""
    uid  = session['user_id']
    name = session['name']
    data = request.json or {}
    existing = data.get('token', '').strip()

    # ── 예매 오픈 시간 체크 (관리자는 제한 없음) ──
    if session.get('role') != 'admin':
        open_time_str = get_config_value('booking_open_time', '')
        if open_time_str:
            try:
                open_dt = datetime.datetime.strptime(open_time_str, '%Y-%m-%d %H:%M:%S')
                # 서버 시간대와 무관하게 한국 시간(KST=UTC+9) 기준으로 비교
                KST = datetime.timezone(datetime.timedelta(hours=9))
                now_dt = datetime.datetime.now(KST).replace(tzinfo=None)
                if now_dt < open_dt:
                    return jsonify({
                        'error': '예매가 아직 시작되지 않았습니다.',
                        'open_time': open_time_str,
                        'open_time_display': open_dt.strftime('%Y년 %m월 %d일 %p %I:%M').replace(
                            'AM', '오전').replace('PM', '오후')
                    }), 403
            except ValueError:
                pass  # 잘못된 형식이면 무시

    with _ql:
        # ① 기존 토큰 복원 (새로고침/재접속)
        if existing:
            if existing in _active:
                v = _active[existing]
                v['heartbeat'] = time.time()
                remaining = max(0, int(ACTIVE_TIMEOUT - (time.time() - v['entered_at'])))
                return jsonify({'status': 'active', 'token': existing,
                                'remaining': remaining, 'active_count': len(_active)})
            if existing in _waiting:
                keys = list(_waiting.keys())
                pos  = keys.index(existing) + 1
                eta  = int((pos / max(len(_active), 1)) * ACTIVE_TIMEOUT / 60)
                return jsonify({'status': 'waiting', 'token': existing, 'position': pos,
                                'active_count': len(_active), 'waiting_count': len(_waiting),
                                'eta_min': max(1, eta)})

        # ① -b 같은 uid가 이미 _active에 있으면 해당 토큰 재사용
        # (뒤로가기·탭 닫기 등으로 sendQueueLeave가 누락됐을 때 토큰 중복 방지)
        for tok, info in _active.items():
            if info['uid'] == uid:
                info['heartbeat'] = time.time()
                remaining = max(0, int(ACTIVE_TIMEOUT - (time.time() - info['entered_at'])))
                return jsonify({'status': 'active', 'token': tok,
                                'remaining': remaining, 'active_count': len(_active)})

        # ② 새 토큰 발급
        token = str(uuid.uuid4())
        if len(_active) < MAX_ACTIVE:
            _active[token] = {'uid': uid, 'name': name,
                               'entered_at': time.time(), 'heartbeat': time.time()}
            return jsonify({'status': 'active', 'token': token,
                            'remaining': ACTIVE_TIMEOUT, 'active_count': len(_active)})
        else:
            _waiting[token] = {'uid': uid, 'name': name, 'joined_at': time.time()}
            pos = len(_waiting)
            eta = int((pos / MAX_ACTIVE) * ACTIVE_TIMEOUT / 60)
            return jsonify({'status': 'waiting', 'token': token, 'position': pos,
                            'active_count': len(_active), 'waiting_count': len(_waiting),
                            'eta_min': max(1, eta)})

@app.route('/api/queue/status')
@login_required
def queue_status():
    """대기열 상태 폴링"""
    token = request.args.get('token', '')
    with _ql:
        if token in _active:
            v = _active[token]
            remaining = max(0, int(ACTIVE_TIMEOUT - (time.time() - v['entered_at'])))
            return jsonify({'status': 'active', 'remaining': remaining,
                            'active_count': len(_active)})
        if token in _waiting:
            keys = list(_waiting.keys())
            pos  = keys.index(token) + 1
            eta  = int((pos / max(len(_active), 1)) * ACTIVE_TIMEOUT / 60)
            return jsonify({'status': 'waiting', 'position': pos,
                            'active_count': len(_active),
                            'waiting_count': len(_waiting),
                            'eta_min': max(1, eta)})
    return jsonify({'status': 'expired'})

@app.route('/api/queue/heartbeat', methods=['POST'])
@csrf_protect
@login_required
def queue_heartbeat():
    """하트비트 갱신 (30초마다 클라이언트가 호출)"""
    token = (request.json or {}).get('token', '')
    with _ql:
        if token in _active:
            _active[token]['heartbeat'] = time.time()
            v = _active[token]
            remaining = max(0, int(ACTIVE_TIMEOUT - (time.time() - v['entered_at'])))
            return jsonify({'ok': True, 'remaining': remaining})
    return jsonify({'ok': False, 'reason': 'not_active'})

@app.route('/api/queue/leave', methods=['POST'])
@login_required
def queue_leave():
    """자발적 퇴장 (브라우저 종료, 뒤로가기 등)"""
    token = (request.json or {}).get('token', '')
    kicked = False
    with _ql:
        if token in _active:
            _queue_kick(token)   # 메모리만 (_ql 내부)
            kicked = True
        elif token in _waiting:
            _waiting.pop(token, None)
    # Perf-②: _ql 해제 후 DB 작업
    if kicked:
        _release_seat_lock_db(token)
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# API - 좌석 선점 Lock
# ──────────────────────────────────────────────
@app.route('/api/seats/lock', methods=['POST'])
@csrf_protect
@login_required
def seat_lock():
    """좌석 선점 (원자적 UPDATE)"""
    data    = request.json or {}
    seat_id = data.get('seat_id', '')
    token   = data.get('token', '')

    with _ql:
        if token not in _active:
            return jsonify({'error': '입장 권한이 없습니다. 대기열을 다시 확인해주세요.'}), 403

    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db()
    cur  = conn.execute(
        "UPDATE seats SET status='locked', locked_by=?, locked_at=? "
        "WHERE id=? AND (status='available')",
        (token, now_str, seat_id))
    conn.commit()

    if cur.rowcount == 0:
        row = conn.execute("SELECT status, locked_by FROM seats WHERE id=?", (seat_id,)).fetchone()
        conn.close()
        if row and row['locked_by'] == token:
            return jsonify({'ok': True, 'mine': True})
        return jsonify({'error': '이미 다른 분이 선택 중인 좌석입니다.'}), 409

    conn.close()
    return jsonify({'ok': True, 'mine': True})

@app.route('/api/seats/unlock', methods=['POST'])
@csrf_protect
@login_required
def seat_unlock():
    """좌석 선점 해제 (내가 lock한 것만)"""
    data    = request.json or {}
    seat_id = data.get('seat_id', '')
    token   = data.get('token', '')
    conn = get_db()
    conn.execute(
        "UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL "
        "WHERE id=? AND locked_by=? AND status='locked'",
        (seat_id, token))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/queue/info')
def queue_info():
    """공개 대기열 현황 (로그인 불필요 - 메인페이지 표시용)"""
    with _ql:
        return jsonify({'active': len(_active), 'waiting': len(_waiting),
                        'max_active': MAX_ACTIVE})

@app.route('/mypage')
def mypage():
    if 'user_id' not in session:
        return render_template('login.html', redirect='/mypage')
    return render_template('mypage.html')

@app.route('/admin/')
@app.route('/admin')
def admin_dashboard():
    if session.get('role') != 'admin':
        return render_template('login.html', redirect='/admin/')
    return render_template('admin/dashboard.html')

@app.route('/admin/members')
def admin_members_page():
    if session.get('role') != 'admin':
        return render_template('login.html', redirect='/admin/members')
    return render_template('admin/members.html')

@app.route('/admin/reservations')
def admin_reservations_page():
    if session.get('role') != 'admin':
        return render_template('login.html', redirect='/admin/reservations')
    return render_template('admin/reservations.html')

@app.route('/admin/seats')
def admin_seats_page():
    if session.get('role') != 'admin':
        return render_template('login.html', redirect='/admin/seats')
    return render_template('admin/seats.html')

# ──────────────────────────────────────────────
# API - 인증
# ──────────────────────────────────────────────
@app.route('/api/register', methods=['POST'])
def register():
    d = request.get_json()
    username = d.get('username','').strip()
    password = d.get('password','').strip()
    name     = d.get('name','').strip()
    phone    = d.get('phone','').strip()
    church   = d.get('church','').strip()
    position = d.get('position','').strip()

    if not all([username, password, name, phone, church, position]):
        return jsonify({'error': '모든 항목을 입력해주세요.'}), 400
    if len(username) < 4 or len(username) > 20:
        return jsonify({'error': '아이디는 4~20자여야 합니다.'}), 400
    if len(password) < 8:
        return jsonify({'error': '비밀번호는 8자 이상이어야 합니다.'}), 400

    conn = get_db()
    try:
        conn.execute('''INSERT INTO users
            (username, password, name, phone, church, position)
            VALUES (?,?,?,?,?,?)''',
            (username, hash_password(password), name, phone, church, position))
        conn.commit()
        return jsonify({'ok': True, 'message': '회원가입이 완료되었습니다.'})
    except sqlite3.IntegrityError:
        return jsonify({'error': '이미 사용 중인 아이디입니다.'}), 409
    finally:
        conn.close()

@app.route('/api/login', methods=['POST'])
@limiter.limit('10 per minute')   # 같은 IP에서 1분에 10회 초과 시 429
def login():
    d = request.get_json()
    username = d.get('username','').strip()
    password = d.get('password','').strip()

    conn = get_db()
    user = conn.execute(
        'SELECT * FROM users WHERE username=? AND is_active=1',
        (username,)
    ).fetchone()

    if not user or not verify_password(user['password'], password):
        conn.close()
        return jsonify({'error': '아이디 또는 비밀번호가 올바르지 않습니다.'}), 401

    # 구 SHA256 단순 해시 → PBKDF2 자동 마이그레이션 (로그인 성공 시 1회)
    if _SHA256_RE.match(user['password']):
        conn.execute('UPDATE users SET password=? WHERE id=?',
                     (hash_password(password), user['id']))
        conn.commit()
    conn.close()

    session['user_id']   = user['id']
    session['username']  = user['username']
    session['name']      = user['name']
    session['role']      = user['role']
    return jsonify({'ok': True, 'role': user['role'], 'name': user['name']})

# ── 429 Rate-Limit 에러 → JSON 응답 (HTML 대신) ─────────────────────
@app.errorhandler(429)
def too_many_requests(e):
    return jsonify({'error': '요청이 너무 많습니다. 잠시 후 다시 시도해 주세요.'}), 429

@app.route('/api/logout', methods=['POST'])
@csrf_protect
def logout():
    session.clear()
    return jsonify({'ok': True})


@app.route('/api/verify-user', methods=['POST'])
@csrf_protect
def verify_user():
    """비밀번호 변경 전 본인확인: 아이디+이름+전화번호+교회명 일치 여부 검증"""
    d = request.get_json() or {}
    username = d.get('username', '').strip()
    name     = d.get('name', '').strip()
    phone    = d.get('phone', '').strip()
    church   = d.get('church', '').strip()
    if not all([username, name, phone, church]):
        return jsonify({'error': '모든 항목을 입력해 주세요.'}), 400
    conn = get_db()
    user = conn.execute(
        'SELECT id FROM users WHERE username=? AND name=? AND phone=? AND church=?',
        (username, name, phone, church)
    ).fetchone()
    conn.close()
    if not user:
        return jsonify({'error': '등록된 회원이 없습니다.'}), 404
    return jsonify({'ok': True})

@app.route('/api/reset-password', methods=['POST'])
@csrf_protect
def reset_password():
    """본인확인 후 비밀번호 재설정"""
    d = request.get_json() or {}
    username    = d.get('username', '').strip()
    name        = d.get('name', '').strip()
    phone       = d.get('phone', '').strip()
    church      = d.get('church', '').strip()
    new_password = d.get('new_password', '').strip()
    if not all([username, name, phone, church, new_password]):
        return jsonify({'error': '모든 항목을 입력해 주세요.'}), 400
    if len(new_password) < 8:
        return jsonify({'error': '비밀번호는 8자 이상이어야 합니다.'}), 400
    conn = get_db()
    user = conn.execute(
        'SELECT id FROM users WHERE username=? AND name=? AND phone=? AND church=?',
        (username, name, phone, church)
    ).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': '등록된 회원이 없습니다.'}), 404
    conn.execute('UPDATE users SET password=? WHERE id=?',
                 (hash_password(new_password), user['id']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/me')
def me():
    if 'user_id' not in session:
        return jsonify({'logged_in': False})
    return jsonify({
        'logged_in': True,
        'user_id':  session['user_id'],
        'username': session['username'],
        'name':     session['name'],
        'role':     session['role'],
        'is_admin': session.get('role') == 'admin'
    })

@app.route('/api/refund-account', methods=['GET'])
@login_required
def get_refund_account():
    """환불 계좌 조회"""
    conn = get_db()
    row = conn.execute(
        'SELECT name, refund_bank, refund_account, refund_holder FROM users WHERE id=?',
        (session['user_id'],)
    ).fetchone()
    conn.close()
    return jsonify({
        'name':           row['name'],
        'refund_bank':    row['refund_bank']    or '',
        'refund_account': row['refund_account'] or '',
        'refund_holder':  row['refund_holder']  or row['name'],
    })

@app.route('/api/refund-account', methods=['POST'])
@csrf_protect
@login_required
def save_refund_account():
    """환불 계좌 저장"""
    data = request.json or {}
    bank    = data.get('refund_bank',    '').strip()
    account = data.get('refund_account', '').strip()
    holder  = data.get('refund_holder',  '').strip()
    if not bank or not account or not holder:
        return jsonify({'error': '은행명, 계좌번호, 예금주를 모두 입력해주세요.'}), 400
    conn = get_db()
    conn.execute(
        'UPDATE users SET refund_bank=?, refund_account=?, refund_holder=? WHERE id=?',
        (bank, account, holder, session['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/check-username')
def check_username():
    username = request.args.get('username','').strip()
    conn = get_db()
    exists = conn.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone()
    conn.close()
    return jsonify({'available': not exists})

# ──────────────────────────────────────────────
# API - 좌석
# ──────────────────────────────────────────────
@app.route('/api/seats/<int:floor>')
def get_seats(floor):
    conn = get_db()
    rows = conn.execute('''
        SELECT s.id, s.section, s.row_num, s.col_num, s.seat_no, s.status,
               s.locked_by,
               r.status as res_status, r.user_id as res_user
        FROM seats s
        LEFT JOIN reservations r ON s.id = r.seat_id AND r.status = 'confirmed'
        WHERE s.floor = ?
    ''', (floor,)).fetchall()
    conn.close()

    seats = {}
    my_id    = session.get('user_id')
    my_token = request.args.get('token', '')   # 클라이언트가 쿼리로 전달

    for row in rows:
        if row['res_status'] == 'confirmed':
            disp = 'mine' if row['res_user'] == my_id else 'reserved'
        elif row['status'] == 'locked':
            # 내가 lock한 좌석 → 'my_lock' (선택됨 표시)
            disp = 'my_lock' if (my_token and row['locked_by'] == my_token) else 'locked'
        else:
            disp = row['status']
        seats[row['id']] = {
            'section': row['section'],
            'row':     row['row_num'],
            'col':     row['col_num'],
            'seat_no': row['seat_no'],
            'status':  disp,
            'floor':   floor   # 층별 금액 계산에 필요
        }

    # 현재 로그인 사용자의 기존 확정 예매 수 (전체 층 기준)
    reserved_count = 0
    if my_id:
        conn2 = get_db()
        reserved_count = conn2.execute(
            'SELECT COUNT(*) FROM reservations WHERE user_id=? AND status="confirmed"',
            (my_id,)
        ).fetchone()[0]
        conn2.close()

    section_info = {
        sec: {'label':           cfg['label'],
              'floor':           cfg['floor'],
              'rows':            cfg['rows'],
              'row_seats':       cfg['row_seats'],
              'align':           cfg.get('align', 'center'),
              'disability_cols': cfg.get('disability_cols', 0),
              'aisle_after_row': cfg.get('aisle_after_row', 0)}   # 복도 위치 (0=없음)
        for sec, cfg in SECTIONS.items()
        if cfg['floor'] == floor
    }
    is_admin = session.get('role') == 'admin'
    return jsonify({'seats': seats, 'sections': section_info,
                    'reserved_count': reserved_count, 'max_seats': 10,
                    'is_admin': is_admin})

@app.route('/api/stats/seats')
def seat_stats():
    conn = get_db()
    blocked  = conn.execute('SELECT COUNT(*) FROM seats WHERE status="blocked"').fetchone()[0]
    total    = conn.execute('SELECT COUNT(*) FROM seats').fetchone()[0] - blocked
    reserved = conn.execute('SELECT COUNT(*) FROM reservations WHERE status="confirmed"').fetchone()[0]
    conn.close()
    return jsonify({'total': total, 'reserved': reserved, 'blocked': blocked,
                    'available': total - reserved})

@app.route('/api/booking-status')
def booking_status():
    """예매 오픈 여부 + 서버 현재 시각 (공개 API, 로그인 불필요)"""
    open_time_str = get_config_value('booking_open_time', '')
    now = datetime.datetime.now()
    is_open = True  # 오픈 시간 미설정 시 기본 활성화
    if open_time_str:
        try:
            open_dt = datetime.datetime.strptime(open_time_str, '%Y-%m-%d %H:%M:%S')
            is_open = now >= open_dt
        except ValueError:
            pass
    price1 = int(get_config_value('ticket_price_floor1', '10000') or 10000)
    price2 = int(get_config_value('ticket_price_floor2', '10000') or 10000)
    return jsonify({
        'is_open':    is_open,
        'open_time':  open_time_str,          # '' 이면 미설정
        'server_time': now.strftime('%Y-%m-%d %H:%M:%S'),
        'ticket_price_floor1': price1,
        'ticket_price_floor2': price2,
    })

# ──────────────────────────────────────────────
# API - 예매
# ──────────────────────────────────────────────
@app.route('/api/reservations', methods=['POST'])
@csrf_protect
@login_required
def create_reservation():
    d = request.get_json()
    seat_ids = d.get('seat_ids', [])

    if not seat_ids:
        return jsonify({'error': '좌석을 선택해주세요.'}), 400

    is_admin  = session.get('role') == 'admin'
    MAX_PER   = 10  # 일반 회원 최대 예매 좌석 수

    if not is_admin and len(seat_ids) > MAX_PER:
        return jsonify({'error': f'1회 최대 {MAX_PER}석까지 예매 가능합니다.'}), 400

    user_id = session['user_id']

    # Perf-③A: 세마포어로 동시 처리 수 제한 (5초 대기 후 포기 → 503)
    if not _reservation_sem.acquire(blocking=True, timeout=5):
        return jsonify({'error': '예매 요청이 많습니다. 잠시 후 다시 시도해주세요.'}), 503

    # isolation_level=None: 수동 트랜잭션 관리 (BEGIN IMMEDIATE 사용을 위해 필요)
    conn = sqlite3.connect(DATABASE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    conn.execute('PRAGMA journal_mode = WAL')
    conn.isolation_level = None  # autocommit 모드 → 명시적 트랜잭션 사용

    def rollback_and_close():
        try:
            conn.execute('ROLLBACK')
        except Exception:
            pass
        conn.close()
        _reservation_sem.release()  # Perf-③: 세마포어 반드시 해제

    try:
        # Perf-③B: BEGIN IMMEDIATE 재시도 (최대 3회, 지수 백오프 50ms→100ms)
        # 동시 예매 충돌 시 즉시 503 대신 짧게 대기 후 재시도
        for _attempt in range(3):
            try:
                conn.execute('BEGIN IMMEDIATE')
                break
            except sqlite3.OperationalError as _e:
                if 'locked' in str(_e).lower() and _attempt < 2:
                    time.sleep(0.08 * (2 ** _attempt))  # 80ms, 160ms (스레드 32 기준)
                    continue
                conn.close()
                _reservation_sem.release()
                return jsonify({'error': '일시적으로 처리 중입니다. 잠시 후 다시 시도해주세요.'}), 503

        existing = conn.execute(
            'SELECT COUNT(*) FROM reservations WHERE user_id=? AND status="confirmed"',
            (user_id,)
        ).fetchone()[0]
        if not is_admin and existing + len(seat_ids) > MAX_PER:
            rollback_and_close()
            return jsonify({'error': f'이미 {existing}석 예매됨. 총 {MAX_PER}석 초과 불가.'}), 400

        created = []
        q_token = d.get('queue_token', '')

        for sid in seat_ids:
            # available 또는 내가 lock한 locked 좌석만 예매 가능
            seat = conn.execute(
                "SELECT * FROM seats WHERE id=? AND (status='available' OR "
                "(status='locked' AND locked_by=?))", (sid, q_token)
            ).fetchone()
            if not seat:
                rollback_and_close()
                return jsonify({'error': f'{sid} 좌석은 예매할 수 없습니다. (다른 분이 선택 중이거나 이미 예매됨)'}), 409

            already = conn.execute(
                'SELECT id FROM reservations WHERE seat_id=? AND status="confirmed"', (sid,)
            ).fetchone()
            if already:
                rollback_and_close()
                return jsonify({'error': f'{sid} 좌석은 이미 예매되었습니다.'}), 409

            # 예매 확정 시 reserved 상태로 전환 (lock 해제)
            conn.execute(
                "UPDATE seats SET status='reserved', locked_by=NULL, locked_at=NULL WHERE id=?",
                (sid,))
            code = gen_res_code(conn)
            conn.execute('''INSERT INTO reservations
                (res_code, user_id, seat_id) VALUES (?,?,?)''',
                (code, user_id, sid))
            created.append({'res_code': code, 'seat_id': sid})

        conn.execute('COMMIT')
        conn.close()
        _reservation_sem.release()  # Perf-③: 정상 완료 후 세마포어 해제
        return jsonify({'ok': True, 'reservations': created,
                        'message': f'{len(created)}석 예매가 완료되었습니다.'})
    except sqlite3.OperationalError as e:
        rollback_and_close()
        if 'locked' in str(e).lower():
            return jsonify({'error': '일시적으로 처리 중입니다. 잠시 후 다시 시도해주세요.'}), 503
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        rollback_and_close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/reservations/my')
@login_required
def my_reservations():
    conn = get_db()
    rows = conn.execute('''
        SELECT r.id, r.res_code, r.status, r.reserved_at, r.cancelled_at,
               r.payment_status, r.payment_at,
               s.section, s.seat_no, s.floor
        FROM reservations r
        JOIN seats s ON r.seat_id = s.id
        WHERE r.user_id = ?
        ORDER BY r.reserved_at DESC
    ''', (session['user_id'],)).fetchall()
    conn.close()
    return jsonify({'reservations': [dict(r) for r in rows]})

@app.route('/api/reservations/<int:res_id>', methods=['DELETE'])
@csrf_protect
@login_required
def cancel_reservation(res_id):
    conn = get_db()
    res = conn.execute(
        'SELECT * FROM reservations WHERE id=? AND user_id=?',
        (res_id, session['user_id'])
    ).fetchone()
    if not res:
        conn.close()
        return jsonify({'error': '예매 내역을 찾을 수 없습니다.'}), 404
    if res['status'] != 'confirmed':
        conn.close()
        return jsonify({'error': '이미 취소된 예매입니다.'}), 400

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Bug-1 수정: 취소 시 해당 좌석을 available 로 복원 (reservations 변경과 같은 트랜잭션)
    conn.execute(
        "UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL WHERE id=?",
        (res['seat_id'],)
    )
    conn.execute(
        'UPDATE reservations SET status="cancelled", cancelled_at=? WHERE id=?',
        (now, res_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'message': '예매가 취소되었습니다.'})

# ──────────────────────────────────────────────
# 티켓 출력
# ──────────────────────────────────────────────
@app.route('/print/tickets/all')
@login_required
def print_tickets_all():
    uid = session['user_id']
    conn = get_db()
    tickets_rows = conn.execute('''
        SELECT r.res_code, r.reserved_at, r.payment_status,
               s.floor, s.section, s.seat_no
        FROM reservations r
        JOIN seats s ON r.seat_id = s.id
        WHERE r.user_id = ? AND r.status = 'confirmed'
        ORDER BY s.floor, s.section, CAST(s.seat_no AS INTEGER)
    ''', (uid,)).fetchall()
    user_row = conn.execute(
        'SELECT name, phone, church, position FROM users WHERE id=?', (uid,)
    ).fetchone()
    conn.close()

    if not tickets_rows:
        return '<h3 style="padding:40px;font-family:sans-serif">예매된 좌석이 없습니다.</h3>'

    floor_map = {1: '1층', 2: '2층'}
    tickets = [{
        'res_code':       r['res_code'],
        'reserved_at':    r['reserved_at'],
        'payment_status': r['payment_status'],
        'floor_str':      floor_map.get(r['floor'], str(r['floor']) + '층'),
        'section':        r['section'],
        'seat_no':        r['seat_no'],
    } for r in tickets_rows]

    return render_template('tickets_all_print.html',
        tickets         = tickets,
        name            = user_row['name'],
        phone           = user_row['phone'],
        church          = user_row['church'],
        position        = user_row['position'],
        total           = len(tickets),
        is_admin_ticket = (session.get('role') == 'admin')
    )

@app.route('/print/tickets/bulk')
@login_required
def print_tickets_bulk():
    """선택한 예매번호들만 모아서 출력 (마이페이지 체크박스 선택)"""
    codes_param = request.args.get('codes', '')
    codes = [c.strip() for c in codes_param.split(',') if c.strip()]
    if not codes:
        return '<h3 style="padding:40px;font-family:sans-serif">선택된 티켓이 없습니다.</h3>'

    uid  = session['user_id']
    conn = get_db()
    placeholders = ','.join(['?'] * len(codes))
    tickets_rows = conn.execute(f'''
        SELECT r.res_code, r.reserved_at, r.payment_status,
               s.floor, s.section, s.seat_no
        FROM reservations r
        JOIN seats s ON r.seat_id = s.id
        WHERE r.user_id = ? AND r.status = 'confirmed'
          AND r.res_code IN ({placeholders})
        ORDER BY s.floor, s.section, CAST(s.seat_no AS INTEGER)
    ''', (uid, *codes)).fetchall()
    user_row = conn.execute(
        'SELECT name, phone, church, position FROM users WHERE id=?', (uid,)
    ).fetchone()
    conn.close()

    if not tickets_rows:
        return '<h3 style="padding:40px;font-family:sans-serif">출력할 티켓이 없습니다.</h3>'

    floor_map = {1: '1층', 2: '2층'}
    tickets = [{
        'res_code':       r['res_code'],
        'reserved_at':    r['reserved_at'],
        'payment_status': r['payment_status'],
        'floor_str':      floor_map.get(r['floor'], str(r['floor']) + '층'),
        'section':        r['section'],
        'seat_no':        r['seat_no'],
    } for r in tickets_rows]

    return render_template('tickets_all_print.html',
        tickets         = tickets,
        name            = user_row['name'],
        phone           = user_row['phone'],
        church          = user_row['church'],
        position        = user_row['position'],
        total           = len(tickets),
        is_admin_ticket = (session.get('role') == 'admin')
    )

@app.route('/print/ticket/<res_code>')
@login_required
def print_ticket(res_code):
    conn = get_db()
    row = conn.execute('''
        SELECT r.res_code, r.reserved_at, r.status, r.payment_status, r.payment_at,
               s.floor, s.section, s.seat_no,
               u.name, u.phone, u.church, u.position, u.id as uid, u.role
        FROM reservations r
        JOIN seats s ON r.seat_id = s.id
        JOIN users u ON r.user_id = u.id
        WHERE r.res_code = ? AND r.status = 'confirmed'
    ''', (res_code,)).fetchone()
    conn.close()

    if not row:
        return '<h3 style="padding:40px;color:red">예매 정보를 찾을 수 없습니다.</h3>', 404

    # 본인 또는 관리자만 접근 가능
    if row['uid'] != session.get('user_id') and session.get('role') != 'admin':
        return '<h3 style="padding:40px;color:red">접근 권한이 없습니다.</h3>', 403

    floor_map = {1: '1층', 2: '2층'}
    return render_template('ticket_print.html',
        res_code       = row['res_code'],
        reserved_at    = row['reserved_at'],
        floor_str      = floor_map.get(row['floor'], str(row['floor']) + '층'),
        section        = row['section'],
        seat_no        = row['seat_no'],
        name           = row['name'],
        phone          = row['phone'],
        church         = row['church'],
        position       = row['position'],
        payment_status = row['payment_status'],
        payment_at     = row['payment_at'],
        is_admin_ticket = (row['role'] == 'admin')
    )

# ──────────────────────────────────────────────
# API - 관리자
# ──────────────────────────────────────────────
@app.route('/api/admin/stats')
@login_required
@admin_required
def admin_stats():
    conn = get_db()
    total_users   = conn.execute('SELECT COUNT(*) FROM users WHERE role="user"').fetchone()[0]
    blocked       = conn.execute('SELECT COUNT(*) FROM seats WHERE status="blocked"').fetchone()[0]
    total_seats   = conn.execute('SELECT COUNT(*) FROM seats').fetchone()[0] - blocked
    confirmed     = conn.execute('SELECT COUNT(*) FROM reservations WHERE status="confirmed"').fetchone()[0]
    cancelled     = conn.execute('SELECT COUNT(*) FROM reservations WHERE status!="confirmed"').fetchone()[0]
    paid          = conn.execute('SELECT COUNT(*) FROM reservations WHERE status="confirmed" AND payment_status="paid"').fetchone()[0]
    unpaid        = confirmed - paid
    conn.close()
    return jsonify({
        'total_users': total_users,
        'total_seats': total_seats,
        'confirmed':   confirmed,
        'cancelled':   cancelled,
        'blocked':     blocked,
        'available':   total_seats - confirmed,
        'paid':        paid,
        'unpaid':      unpaid
    })

@app.route('/api/admin/members')
@login_required
@admin_required
def admin_get_members():
    search   = request.args.get('search','')
    church   = request.args.get('church','')
    position = request.args.get('position','')
    page     = int(request.args.get('page', 1))
    per_page = 20

    conn = get_db()
    where, params = ['role != "admin"'], []
    if search:
        where.append('(name LIKE ? OR username LIKE ?)')
        params += [f'%{search}%', f'%{search}%']
    if church:
        where.append('church LIKE ?')
        params.append(f'%{church}%')
    if position:
        where.append('position = ?')
        params.append(position)

    w_sql = ' AND '.join(where)
    total = conn.execute(f'SELECT COUNT(*) FROM users WHERE {w_sql}', params).fetchone()[0]
    rows  = conn.execute(
        f'''SELECT id, username, name, phone, church, position,
                   is_active, created_at,
                   refund_bank, refund_account, refund_holder
            FROM users WHERE {w_sql}
            ORDER BY id DESC LIMIT ? OFFSET ?''',
        params + [per_page, (page-1)*per_page]
    ).fetchall()
    conn.close()
    return jsonify({'members': [dict(r) for r in rows], 'total': total, 'page': page})

@app.route('/api/admin/members/<int:uid>', methods=['PUT'])
@csrf_protect
@login_required
@admin_required
def admin_update_member(uid):
    d = request.get_json()
    conn = get_db()
    conn.execute('''UPDATE users SET name=?, phone=?, church=?, position=?, is_active=?
        WHERE id=?''',
        (d['name'], d['phone'], d['church'], d['position'],
         1 if d.get('is_active') else 0, uid))

    # 비밀번호 강제 변경 (입력된 경우에만)
    new_pw = d.get('new_password', '').strip()
    if new_pw:
        if len(new_pw) < 8:
            conn.close()
            return jsonify({'error': '비밀번호는 8자 이상이어야 합니다.'}), 400
        conn.execute('UPDATE users SET password=? WHERE id=?',
                     (hash_password(new_pw), uid))

    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/members/<int:uid>', methods=['DELETE'])
@csrf_protect
@login_required
@admin_required
def admin_delete_member(uid):
    conn = get_db()

    # Bug-3 수정: reservations 취소 전에 seat_id 목록을 먼저 조회 → 좌석 복원
    confirmed_seats = conn.execute(
        'SELECT seat_id FROM reservations WHERE user_id=? AND status="confirmed"', (uid,)
    ).fetchall()
    for seat_row in confirmed_seats:
        conn.execute(
            "UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL WHERE id=?",
            (seat_row['seat_id'],)
        )

    conn.execute('UPDATE reservations SET status="admin_cancelled" WHERE user_id=? AND status="confirmed"', (uid,))
    conn.execute('UPDATE users SET is_active=0 WHERE id=?', (uid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/reservations')
@login_required
@admin_required
def admin_get_reservations():
    search   = request.args.get('search','')
    status   = request.args.get('status','')
    floor    = request.args.get('floor','')
    section  = request.args.get('section','')
    page     = int(request.args.get('page', 1))
    per_page = 30

    conn = get_db()
    where, params = ['1=1'], []
    if search:
        where.append('(u.name LIKE ? OR u.church LIKE ? OR u.username LIKE ?)')
        params += [f'%{search}%', f'%{search}%', f'%{search}%']
    if status:
        where.append('r.status = ?')
        params.append(status)
    if floor:
        where.append('s.floor = ?')
        params.append(int(floor))
    if section:
        where.append('s.section = ?')
        params.append(section)

    w_sql = ' AND '.join(where)
    total = conn.execute(f'''SELECT COUNT(*) FROM reservations r
        JOIN users u ON r.user_id=u.id JOIN seats s ON r.seat_id=s.id
        WHERE {w_sql}''', params).fetchone()[0]

    rows = conn.execute(f'''
        SELECT r.id, r.res_code, r.status, r.reserved_at, r.cancelled_at, r.memo,
               r.payment_status, r.payment_at,
               u.id as uid, u.name, u.username, u.phone, u.church, u.position,
               s.id as seat_id, s.floor, s.section, s.seat_no
        FROM reservations r
        JOIN users u ON r.user_id=u.id
        JOIN seats s ON r.seat_id=s.id
        WHERE {w_sql}
        ORDER BY r.id DESC LIMIT ? OFFSET ?
    ''', params + [per_page, (page-1)*per_page]).fetchall()
    conn.close()
    return jsonify({'reservations': [dict(r) for r in rows], 'total': total, 'page': page})

@app.route('/api/admin/reservations/<int:res_id>', methods=['PUT'])
@csrf_protect
@login_required
@admin_required
def admin_update_reservation(res_id):
    d = request.get_json()
    conn = get_db()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_status = d.get('status', 'confirmed')
    cancelled_at = now if new_status != 'confirmed' else None

    # Bug-2 수정: 취소 계열 상태로 변경 시 해당 좌석을 available 로 복원
    if new_status != 'confirmed':
        res_row = conn.execute(
            'SELECT seat_id FROM reservations WHERE id=? AND status="confirmed"', (res_id,)
        ).fetchone()
        if res_row:
            conn.execute(
                "UPDATE seats SET status='available', locked_by=NULL, locked_at=NULL WHERE id=?",
                (res_row['seat_id'],)
            )

    conn.execute('UPDATE reservations SET status=?, cancelled_at=?, memo=? WHERE id=?',
                 (new_status, cancelled_at, d.get('memo',''), res_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/reservations/<int:res_id>/payment', methods=['PUT'])
@csrf_protect
@login_required
@admin_required
def admin_update_payment(res_id):
    d = request.get_json()
    action = d.get('action')  # 'pay' 또는 'cancel'
    conn = get_db()
    if action == 'pay':
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute(
            'UPDATE reservations SET payment_status="paid", payment_at=? WHERE id=?',
            (now, res_id)
        )
    else:
        conn.execute(
            'UPDATE reservations SET payment_status=NULL, payment_at=NULL WHERE id=?',
            (res_id,)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/admin/reservations/bulk-payment', methods=['POST'])
@csrf_protect
@login_required
@admin_required
def admin_bulk_payment():
    """선택한 예매 건 일괄 입금처리/취소"""
    d      = request.get_json() or {}
    ids    = d.get('ids', [])
    action = d.get('action')  # 'pay' 또는 'cancel'
    if not ids or action not in ('pay', 'cancel'):
        return jsonify({'error': '잘못된 요청입니다.'}), 400
    conn = get_db()
    if action == 'pay':
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for rid in ids:
            conn.execute(
                'UPDATE reservations SET payment_status="paid", payment_at=? WHERE id=?',
                (now, rid)
            )
    else:
        for rid in ids:
            conn.execute(
                'UPDATE reservations SET payment_status=NULL, payment_at=NULL WHERE id=?',
                (rid,)
            )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'processed': len(ids)})

@app.route('/api/admin/reservations/<int:res_id>', methods=['DELETE'])
@csrf_protect
@login_required
@admin_required
def admin_delete_reservation(res_id):
    conn = get_db()
    # 입금완료 건은 삭제 불가 (서버 이중 보호)
    row = conn.execute(
        'SELECT payment_status FROM reservations WHERE id=?', (res_id,)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '예매 내역을 찾을 수 없습니다.'}), 404
    if row['payment_status'] == 'paid':
        conn.close()
        return jsonify({'error': '입금완료 티켓은 삭제할 수 없습니다. 먼저 입금 취소 후 삭제해 주세요.'}), 403
    conn.execute('DELETE FROM reservations WHERE id=?', (res_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/seats/<seat_id>', methods=['PUT'])
@csrf_protect
@login_required
@admin_required
def admin_update_seat(seat_id):
    d = request.get_json()
    conn = get_db()
    conn.execute('UPDATE seats SET status=? WHERE id=?', (d['status'], seat_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/admin/export')
@login_required
@admin_required
def admin_export():
    conn = get_db()
    rows = conn.execute('''
        SELECT r.res_code, r.status, r.reserved_at, r.cancelled_at,
               r.payment_status, r.payment_at, r.memo,
               u.username, u.name, u.phone, u.church, u.position,
               u.refund_bank, u.refund_account, u.refund_holder,
               s.floor, s.section, s.seat_no
        FROM reservations r
        JOIN users u ON r.user_id=u.id
        JOIN seats s ON r.seat_id=s.id
        ORDER BY r.id
    ''').fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '예매현황'

    header_fill = PatternFill('solid', fgColor='1B6CA8')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['예매번호', '예매상태', '예매일시', '취소일시', '입금상태', '입금일시',
               '아이디', '성명', '전화번호', '환불은행', '환불계좌', '예금주', '출석교회', '직분',
               '층', '구역', '좌석번호', '메모']
    col_widths = [22, 12, 20, 20, 12, 20, 14, 10, 16, 14, 18, 12, 20, 10, 6, 10, 10, 30]

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    status_map  = {'confirmed': '정상', 'cancelled': '취소', 'admin_cancelled': '관리자취소'}
    payment_map = {'paid': '입금완료'}
    alt_fill    = PatternFill('solid', fgColor='EBF3FB')
    paid_fill   = PatternFill('solid', fgColor='D1FAE5')  # 입금완료 연초록

    for ri, row in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        pay_status = payment_map.get(row['payment_status'] or '', '미입금')
        vals = [
            row['res_code'], status_map.get(row['status'], row['status']),
            row['reserved_at'], row['cancelled_at'] or '-',
            pay_status, row['payment_at'] or '-',
            row['username'], row['name'], row['phone'],
            row['refund_bank'] or '', row['refund_account'] or '', row['refund_holder'] or '',
            row['church'], row['position'],
            f"{row['floor']}층", row['section'] + ' 구역', row['seat_no'],
            row['memo'] or ''
        ]
        memo_col = len(headers)  # 메모는 마지막 열
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal='left' if ci == memo_col else 'center',
                                       vertical='center', wrap_text=(ci == memo_col))
            cell.border = thin
            if fill:
                cell.fill = fill

    # 입금완료 행 초록 표시
    for ri, row in enumerate(rows, 2):
        if row['payment_status'] == 'paid':
            for ci in range(1, len(headers)+1):
                ws.cell(row=ri, column=ci).fill = paid_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows)+1}'

    ws2 = wb.create_sheet('요약')
    ws2['A1'] = '구분'; ws2['B1'] = '건수'
    for cell in [ws2['A1'], ws2['B1']]:
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin
    summary = [
        ('전체 예매', len(rows)),
        ('정상', sum(1 for r in rows if r['status']=='confirmed')),
        ('취소', sum(1 for r in rows if r['status']!='confirmed')),
    ]
    for ri, (k, v) in enumerate(summary, 2):
        ws2.cell(row=ri, column=1, value=k).border = thin
        ws2.cell(row=ri, column=2, value=v).border = thin

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.date.today().strftime('%Y%m%d')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'예매현황_{today}.xlsx')

@app.route('/api/admin/seats')
@login_required
@admin_required
def admin_get_seats():
    """좌석 현황 조회 (관리자 좌석 관리용)"""
    floor   = request.args.get('floor', type=int)
    section = request.args.get('section', '')

    conn = get_db()
    where, params = ['1=1'], []
    if floor:
        where.append('floor=?')
        params.append(floor)
    if section:
        where.append('section=?')
        params.append(section)

    rows = conn.execute(
        f"SELECT id, floor, section, row_num, col_num, seat_no, status "
        f"FROM seats WHERE {' AND '.join(where)} ORDER BY section, row_num, col_num",
        params
    ).fetchall()
    conn.close()
    return jsonify({'seats': [dict(r) for r in rows]})

@app.route('/api/admin/seats/toggle-block', methods=['POST'])
@csrf_protect
@login_required
@admin_required
def admin_toggle_block_seat():
    """좌석 차단/해제 토글 (available ↔ blocked)"""
    data    = request.json or {}
    seat_id = data.get('seat_id', '').strip()
    if not seat_id:
        return jsonify({'error': '좌석 ID가 필요합니다.'}), 400

    conn = get_db()
    row = conn.execute("SELECT status FROM seats WHERE id=?", (seat_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '존재하지 않는 좌석입니다.'}), 404

    current = row['status']
    if current == 'reserved':
        conn.close()
        return jsonify({'error': '예매된 좌석은 차단할 수 없습니다.'}), 409

    new_status = 'available' if current == 'blocked' else 'blocked'
    conn.execute(
        "UPDATE seats SET status=?, locked_by=NULL, locked_at=NULL WHERE id=?",
        (new_status, seat_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'seat_id': seat_id, 'status': new_status})

@app.route('/api/admin/config', methods=['GET'])
@login_required
@admin_required
def admin_get_config():
    """시스템 설정 조회"""
    conn = get_db()
    rows = conn.execute(
        "SELECT key, value FROM config WHERE key IN "
        "('max_active','booking_open_time','booking_duration',"
        "'ticket_price_floor1','ticket_price_floor2')"
    ).fetchall()
    conn.close()
    cfg = {r['key']: (r['value'] or '') for r in rows}
    cfg.setdefault('max_active',           '20')
    cfg.setdefault('booking_open_time',    '')
    cfg.setdefault('booking_duration',     '180')
    cfg.setdefault('ticket_price_floor1',  '10000')
    cfg.setdefault('ticket_price_floor2',  '10000')
    return jsonify(cfg)

@app.route('/api/admin/config', methods=['POST'])
@login_required
@admin_required
@csrf_protect
def admin_set_config():
    """시스템 설정 저장"""
    data = request.get_json() or {}
    conn = get_db()
    allowed = ('max_active', 'booking_open_time', 'booking_duration',
               'ticket_price_floor1', 'ticket_price_floor2')
    for key in allowed:
        if key in data:
            conn.execute("INSERT OR REPLACE INTO config VALUES (?, ?)", (key, str(data[key])))
    conn.commit()
    conn.close()

    # 메모리 전역 변수 즉시 반영
    global MAX_ACTIVE, ACTIVE_TIMEOUT, SEAT_LOCK_SECS
    try:
        if 'max_active' in data:
            MAX_ACTIVE = int(data['max_active'])
        if 'booking_duration' in data:
            v = int(data['booking_duration'])
            ACTIVE_TIMEOUT = v
            SEAT_LOCK_SECS = v
    except Exception:
        pass

    _invalidate_config_cache()  # Perf-①: 설정 변경 시 캐시 즉시 무효화
    return jsonify({'ok': True})


# ────────────────────────────────────────────────────────────
# 앱 시작
# ────────────────────────────────────────────────────────────
init_db()

# ── 앱 시작 시 SQLite 전역 최적화 (프로세스당 1회) ──────────────
try:
    _boot_conn = sqlite3.connect(DATABASE, timeout=10)
    _boot_conn.execute('PRAGMA journal_mode = WAL')
    _boot_conn.execute('PRAGMA synchronous = NORMAL')
    _boot_conn.execute('PRAGMA cache_size = -32000')   # 32MB 캐시
    _boot_conn.execute('PRAGMA mmap_size = 134217728') # 128MB mmap
    _boot_conn.execute('PRAGMA wal_autocheckpoint = 2000')
    _boot_conn.execute('PRAGMA optimize')              # 쿼리 플래너 통계 갱신
    _boot_conn.commit()
    _boot_conn.close()
except Exception:
    pass

# DB config에서 설정값 로드
try:
    _cfg_max = get_config_value('max_active')
    if _cfg_max:
        MAX_ACTIVE = int(_cfg_max)
    _cfg_dur = get_config_value('booking_duration')
    if _cfg_dur:
        ACTIVE_TIMEOUT = int(_cfg_dur)
        SEAT_LOCK_SECS = int(_cfg_dur)
except Exception:
    pass

if __name__ == "__main__":
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    print("=" * 50)
    print("Server started at http://127.0.0.1:5000")
    print(f"Mode: {'DEBUG (개발)' if debug_mode else 'PRODUCTION (운영)'}")
    print("Admin: admin / Admin1234!")
    print("=" * 50)
    app.run(debug=debug_mode, host="0.0.0.0", port=5000)
